import pandas as pd
from dotenv import load_dotenv
import matplotlib.pyplot as plt
import openpyxl
# from tkinter import Tk
# from tkinter.filedialog import askopenfilename
import seaborn as sns
load_dotenv()

# ask for file input
filename1 = "data/scrimmage_yakker/11_12_2.csv"
filename2 = "data/scrimmage_yakker/11_12_3.csv"
filename3 = "data/scrimmage_yakker/11_12_4.csv"

yakker1 = pd.read_csv(filename1)
yakker2 = pd.read_csv(filename2)
yakker3 = pd.read_csv(filename3)

yakker = pd.concat([yakker1, yakker2, yakker3])

date = yakker['Date'].unique()[0]
date = date.replace('/', '_')

away_team = input('Enter away team: ')

home_team = 'GT'

umpire = input('Enter umpire name: ')

title = f'{away_team}_{home_team}_{date}.xlsx'

yakker = yakker[(yakker['PitchCall'].isin(['BallCalled', 'StrikeCalled']))].reset_index(drop=True) 

yakker['actual_strike'] = (yakker['PlateLocSide'] - 1/6 < 17/24) & (yakker['PlateLocSide'] + 1/6 > -17/24) & (yakker['PlateLocHeight'] - 1/6 < 3) & (yakker['PlateLocHeight'] + 1/6 > 5/4)
yakker['called_strike'] = yakker['PitchCall'] == 'StrikeCalled'

yakker.to_csv('ump_scorecards/yakker.csv')

workbook = openpyxl.Workbook()

worksheet = workbook.active

worksheet.merge_cells('C1:G2')
worksheet['C1'] = 'Umpire Scorecard'
worksheet['C1'].font = openpyxl.styles.Font(size=20, bold=True)
worksheet['C1'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet.merge_cells('D44:F44')
worksheet['D44'] = umpire
worksheet['D44'].font = openpyxl.styles.Font(size=14)
worksheet['D44'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet.merge_cells('D3:F3')
worksheet['D3'] = f'{away_team} @ {home_team}'
worksheet['D3'].font = openpyxl.styles.Font(size=14)
worksheet['D3'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet.merge_cells('D4:F4')
worksheet['D4'] = f'{date.replace("_", "/")}'
worksheet['D4'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

colors = {'BallCalled': 'blue', 'StrikeCalled': 'red'}

plt.scatter(yakker['PlateLocSide'], yakker['PlateLocHeight'], c=yakker['PitchCall'].map(colors), s=40)
plt.plot([-17/24, 17/24], [5/4, 5/4], color='black')
plt.plot([-17/24, 17/24], [3, 3], color='black')
plt.plot([-17/24, -17/24], [5/4, 3], color='black')
plt.plot([17/24, 17/24], [5/4, 3], color='black')
plt.plot([29/24, 29/24], [-1, 6], color='black', linestyle='dotted')
plt.plot([-29/24, -29/24], [-1, 6], color='black', linestyle='dotted')
plt.xlim(-2, 2)
plt.ylim(0, 5)
plt.xticks([])
plt.yticks([])
plt.title('Called Pitches')
plt.gcf().set_size_inches(1.9, 2.8)
plt.savefig('ump_scorecards/called_pitches.png')
plt.close()

img = openpyxl.drawing.image.Image('ump_scorecards/called_pitches.png')
worksheet.add_image(img, 'A6')

edge_calls = yakker[
    (((abs(yakker['PlateLocSide'] - 17/24) < 1/4) | (abs(yakker['PlateLocSide'] + 17/24) < 1/4)) & (yakker['PlateLocHeight'] < 3) & (yakker['PlateLocHeight'] > 5/4)) |
    (((abs(yakker['PlateLocHeight'] - 5/4) < 1/4) | (abs(yakker['PlateLocHeight'] - 3) < 1/4)) & (yakker['PlateLocSide'] < 17/24) & (yakker['PlateLocSide'] > -17/24))
]

plt.scatter(edge_calls['PlateLocSide'], edge_calls['PlateLocHeight'], c=edge_calls['PitchCall'].map(colors), s=40)
plt.plot([-17/24, 17/24], [5/4, 5/4], color='black')
plt.plot([-17/24, 17/24], [3, 3], color='black')
plt.plot([-17/24, -17/24], [5/4, 3], color='black')
plt.plot([17/24, 17/24], [5/4, 3], color='black')
plt.plot([29/24, 29/24], [-1, 6], color='black', linestyle='dotted')
plt.plot([-29/24, -29/24], [-1, 6], color='black', linestyle='dotted')
plt.xlim(-2, 2)
plt.ylim(0, 5)
plt.xticks([])
plt.yticks([])
plt.title('Edge Calls')
plt.gcf().set_size_inches(1.9, 2.8)
plt.savefig('ump_scorecards/edge_calls.png')
plt.close()

img = openpyxl.drawing.image.Image('ump_scorecards/edge_calls.png')
worksheet.add_image(img, 'D6')

called_strikes = yakker[yakker['PitchCall'] == 'StrikeCalled']

# do a density plot of the called strikes
plt.scatter(called_strikes['PlateLocSide'], called_strikes['PlateLocHeight'], s=0)
sns.kdeplot(data=called_strikes, x='PlateLocSide', y='PlateLocHeight', levels = 4, cmap='coolwarm', fill=True, alpha=.5, linewidths=0)
plt.plot([-17/24, 17/24], [5/4, 5/4], color='black')
plt.plot([-17/24, 17/24], [3, 3], color='black')
plt.plot([-17/24, -17/24], [5/4, 3], color='black')
plt.plot([17/24, 17/24], [5/4, 3], color='black')
plt.plot([29/24, 29/24], [-1, 6], color='black', linestyle='dotted')
plt.plot([-29/24, -29/24], [-1, 6], color='black', linestyle='dotted')
plt.xlim(-2, 2)
plt.ylim(0, 5)
plt.xticks([])
plt.yticks([])
plt.xlabel('')
plt.ylabel('')
plt.title('Called Strikes')
plt.gcf().set_size_inches(1.9, 2.8)
plt.savefig('ump_scorecards/density_plot.png')
plt.close()

img = openpyxl.drawing.image.Image('ump_scorecards/density_plot.png')
worksheet.add_image(img, 'G6')

def classify_pitch(pitch):
    # if pd.isna(pitch['SpinAxis']):
    #     return None
    # if pitch['PitcherThrows'] == 'Right':
    #     if 45 <= pitch['SpinAxis'] <= 135:
    #         return 'Curve'
    #     elif 135 <= pitch['SpinAxis'] <= 225:
    #         return 'Rise'
    #     elif 225 <= pitch['SpinAxis'] <= 315:
    #         return 'Screw'
    #     else:
    #         return 'Drop'
    # else:
    #     if 45 <= pitch['SpinAxis'] <= 135:
    #         return 'Screw'
    #     elif 135 <= pitch['SpinAxis'] <= 225:
    #         return 'Rise'
    #     elif 225 <= pitch['SpinAxis'] <= 315:
    #         return 'Curve'
    #     else:
    #         return 'Drop'
    return "Fast"

# yakker['TaggedPitchType'] = yakker.apply(lambda row: classify_pitch(row) if row['BatterTeam'] == 'Georgia tech' else row['TaggedPitchType'], axis=1)
# yakker = yakker.dropna(subset=['TaggedPitchType'])


# yakker['TaggedPitchType'] = yakker['TaggedPitchType'].str.replace('ball', '')
# yakker['TaggedPitchType'] = yakker['TaggedPitchType'].str.replace('up', '')

max_speeds = yakker.groupby('Pitcher')['RelSpeed'].max()

# for i, row in yakker.iterrows():
#     cur_max = max_speeds[row['Pitcher']]
#     if row['RelSpeed'] < cur_max - 10:
#         yakker.at[i, 'TaggedPitchType'] = 'Change'


# get total pitches and number of missed calls by pitch type
pitch_counts = yakker['TaggedPitchType'].value_counts().sort_index()
missed_calls = yakker[yakker['actual_strike'] != yakker['called_strike']]['TaggedPitchType'].value_counts()
accuracy = 1 - (missed_calls / pitch_counts)

for row in range(22, 32):
    worksheet.cell(row = row, column = 5).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
    worksheet.cell(row = row, column = 6).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
    worksheet.cell(row = row, column = 1).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
    worksheet.cell(row = row, column = 9).border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 21, column = 1).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 21, column = 2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 21, column = 3).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 21, column = 4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 24, column = 1).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), left=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 24, column = 2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 24, column = 3).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 24, column = 4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 31, column = 1).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), left=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 31, column = 2).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 31, column = 3).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 31, column = 4).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))

worksheet.merge_cells('A22:D23')
worksheet['A22'] = 'Total Accuracy'
worksheet['A22'].font = openpyxl.styles.Font(bold=True)
worksheet['A22'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['A24'] = 'Pitch'
worksheet['A24'].font = openpyxl.styles.Font(bold=True)
worksheet['A24'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['B24'] = 'Total'
worksheet['B24'].font = openpyxl.styles.Font(bold=True)
worksheet['B24'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['C24'] = 'Misses'
worksheet['C24'].font = openpyxl.styles.Font(bold=True)
worksheet['C24'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['D24'] = 'Accuracy'
worksheet['D24'].font = openpyxl.styles.Font(bold=True)
worksheet['D24'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

for i, pitch in enumerate(pitch_counts.index):
    worksheet[f'A{i + 25}'] = pitch
    worksheet[f'B{i + 25}'] = pitch_counts[pitch]
    worksheet[f'C{i + 25}'] = missed_calls.get(pitch, 0)

    cur_accuracy = accuracy.get(pitch, 1)
    if pd.isna(cur_accuracy):
        cur_accuracy = 1

    worksheet[f'D{i + 25}'] = f'{round(cur_accuracy * 100, 1)}%'
    worksheet[f'D{i + 25}'].alignment = openpyxl.styles.Alignment(horizontal='right')

worksheet['A31'] = 'Total'
worksheet['A31'].font = openpyxl.styles.Font(bold=True)

worksheet['A31'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['B31'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['C31'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['D31'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')

total_called = yakker['PitchCall'].value_counts().sum()
total_missed = yakker[yakker['actual_strike'] != yakker['called_strike']]['PitchCall'].value_counts().sum()
total_accuracy = 1 - (total_missed / total_called)

worksheet['B31'] = total_called
worksheet['C31'] = total_missed
worksheet['D31'] = f'{round(total_accuracy * 100, 1)}%'
worksheet['D31'].alignment = openpyxl.styles.Alignment(horizontal='right')


edge_calls = yakker[
    (((abs(yakker['PlateLocSide'] - 17/24) < 1/4) | (abs(yakker['PlateLocSide'] + 17/24) < 1/4)) & (yakker['PlateLocHeight'] < 3) & (yakker['PlateLocHeight'] > 5/4)) |
    (((abs(yakker['PlateLocHeight'] - 5/4) < 1/4) | (abs(yakker['PlateLocHeight'] - 3) < 1/4)) & (yakker['PlateLocSide'] < 17/24) & (yakker['PlateLocSide'] > -17/24))
]

pitch_counts = edge_calls['TaggedPitchType'].value_counts().sort_index()
missed_calls = edge_calls[edge_calls['actual_strike'] != edge_calls['called_strike']]['TaggedPitchType'].value_counts()
accuracy = 1 - (missed_calls / pitch_counts)

worksheet.cell(row = 21, column = 6).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 21, column = 7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 21, column = 8).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 21, column = 9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 24, column = 6).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 24, column = 7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 24, column = 8).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 24, column = 9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 31, column = 6).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 31, column = 7).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 31, column = 8).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 31, column = 9).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))


worksheet.merge_cells('F22:I23')
worksheet['F22'] = 'Edge Accuracy'
worksheet['F22'].font = openpyxl.styles.Font(bold=True)
worksheet['F22'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['F24'] = 'Pitch'
worksheet['F24'].font = openpyxl.styles.Font(bold=True)
worksheet['F24'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['G24'] = 'Total'
worksheet['G24'].font = openpyxl.styles.Font(bold=True)
worksheet['G24'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['H24'] = 'Misses'
worksheet['H24'].font = openpyxl.styles.Font(bold=True)
worksheet['H24'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['I24'] = 'Accuracy'
worksheet['I24'].font = openpyxl.styles.Font(bold=True)
worksheet['I24'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

for i, pitch in enumerate(pitch_counts.index):
    worksheet[f'F{i + 25}'] = pitch
    worksheet[f'G{i + 25}'] = pitch_counts[pitch]
    worksheet[f'H{i + 25}'] = missed_calls.get(pitch, 0)

    cur_accuracy = accuracy.get(pitch, 1)
    if pd.isna(cur_accuracy):
        cur_accuracy = 1

    worksheet[f'I{i + 25}'] = f'{round(cur_accuracy * 100, 1)}%'
    worksheet[f'I{i + 25}'].alignment = openpyxl.styles.Alignment(horizontal='right')

worksheet['F31'] = 'Total'
worksheet['F31'].font = openpyxl.styles.Font(bold=True)

worksheet['F31'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['G31'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['H31'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['I31'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')

total_called = edge_calls['PitchCall'].value_counts().sum()
total_missed = edge_calls[edge_calls['actual_strike'] != edge_calls['called_strike']]['PitchCall'].value_counts().sum()
total_accuracy = 1 - (total_missed / total_called)

worksheet['G31'] = total_called
worksheet['H31'] = total_missed
worksheet['I31'] = f'{round(total_accuracy * 100, 1)}%'
worksheet['I31'].alignment = openpyxl.styles.Alignment(horizontal='right')

in_zone = yakker[yakker['actual_strike']]

pitch_counts = in_zone['TaggedPitchType'].value_counts().sort_index()
missed_calls = in_zone[in_zone['actual_strike'] != in_zone['called_strike']]['TaggedPitchType'].value_counts()
accuracy = 1 - (missed_calls / pitch_counts)

for row in range(33, 43):
    worksheet.cell(row = row, column = 5).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
    worksheet.cell(row = row, column = 6).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
    worksheet.cell(row = row, column = 1).border = openpyxl.styles.Border(left=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
    worksheet.cell(row = row, column = 9).border = openpyxl.styles.Border(right=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 32, column = 1).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 32, column = 2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 32, column = 3).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 32, column = 4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 34, column = 1).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 34, column = 2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 34, column = 3).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 34, column = 4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 35, column = 1).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), left=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 35, column = 2).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 35, column = 3).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 35, column = 4).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 42, column = 1).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), left=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 42, column = 2).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 42, column = 3).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 42, column = 4).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))

worksheet.merge_cells('A33:D34')
worksheet['A33'] = 'In Zone Accuracy'
worksheet['A33'].font = openpyxl.styles.Font(bold=True)
worksheet['A33'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['A35'] = 'Pitch'
worksheet['A35'].font = openpyxl.styles.Font(bold=True)
worksheet['A35'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['B35'] = 'Total'
worksheet['B35'].font = openpyxl.styles.Font(bold=True)
worksheet['B35'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['C35'] = 'Misses'
worksheet['C35'].font = openpyxl.styles.Font(bold=True)
worksheet['C35'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['D35'] = 'Accuracy'
worksheet['D35'].font = openpyxl.styles.Font(bold=True)
worksheet['D35'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

for i, pitch in enumerate(pitch_counts.index):
    worksheet[f'A{i + 36}'] = pitch
    worksheet[f'B{i + 36}'] = pitch_counts[pitch]
    worksheet[f'C{i + 36}'] = missed_calls.get(pitch, 0)

    cur_accuracy = accuracy.get(pitch, 1)
    if pd.isna(cur_accuracy):
        cur_accuracy = 1

    worksheet[f'D{i + 36}'] = f'{round(cur_accuracy * 100, 1)}%'
    worksheet[f'D{i + 36}'].alignment = openpyxl.styles.Alignment(horizontal='right')

worksheet['A42'] = 'Total'
worksheet['A42'].font = openpyxl.styles.Font(bold=True)

worksheet['A42'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['B42'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['C42'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['D42'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')

total_called = in_zone['PitchCall'].value_counts().sum()
total_missed = in_zone[in_zone['actual_strike'] != in_zone['called_strike']]['PitchCall'].value_counts().sum()
total_accuracy = 1 - (total_missed / total_called)

worksheet['B42'] = total_called
worksheet['C42'] = total_missed
worksheet['D42'] = f'{round(total_accuracy * 100, 1)}%'
worksheet['D42'].alignment = openpyxl.styles.Alignment(horizontal='right')

out_of_zone = yakker[~yakker['actual_strike']]

pitch_counts = out_of_zone['TaggedPitchType'].value_counts().sort_index()
missed_calls = out_of_zone[out_of_zone['actual_strike'] != out_of_zone['called_strike']]['TaggedPitchType'].value_counts()
accuracy = 1 - (missed_calls / pitch_counts)

worksheet.cell(row = 32, column = 6).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 32, column = 7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 32, column = 8).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 32, column = 9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 35, column = 6).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 35, column = 7).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 35, column = 8).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 35, column = 9).border = openpyxl.styles.Border(bottom=openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'), top=openpyxl.styles.Side(style='thin'))

worksheet.cell(row = 42, column = 6).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 42, column = 7).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 42, column = 8).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'))
worksheet.cell(row = 42, column = 9).border = openpyxl.styles.Border(bottom = openpyxl.styles.Side(style='thin'), right=openpyxl.styles.Side(style='thin'))

worksheet.merge_cells('F33:I34')
worksheet['F33'] = 'Out of Zone Accuracy'
worksheet['F33'].font = openpyxl.styles.Font(bold=True)
worksheet['F33'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['F35'] = 'Pitch'
worksheet['F35'].font = openpyxl.styles.Font(bold=True)
worksheet['F35'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['G35'] = 'Total'
worksheet['G35'].font = openpyxl.styles.Font(bold=True)
worksheet['G35'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['H35'] = 'Misses'
worksheet['H35'].font = openpyxl.styles.Font(bold=True)
worksheet['H35'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

worksheet['I35'] = 'Accuracy'
worksheet['I35'].font = openpyxl.styles.Font(bold=True)
worksheet['I35'].alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')

for i, pitch in enumerate(pitch_counts.index):
    worksheet[f'F{i + 36}'] = pitch
    worksheet[f'G{i + 36}'] = pitch_counts[pitch]
    worksheet[f'H{i + 36}'] = missed_calls.get(pitch, 0)

    cur_accuracy = accuracy.get(pitch, 1)
    if pd.isna(cur_accuracy):
        cur_accuracy = 1

    worksheet[f'I{i + 36}'] = f'{round(cur_accuracy * 100, 1)}%'
    worksheet[f'I{i + 36}'].alignment = openpyxl.styles.Alignment(horizontal='right')

worksheet['F42'] = 'Total'
worksheet['F42'].font = openpyxl.styles.Font(bold=True)

worksheet['F42'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['G42'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['H42'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')
worksheet['I42'].fill = openpyxl.styles.PatternFill(start_color='d0d0d0', end_color='d0d0d0', fill_type='solid')

total_called = out_of_zone['PitchCall'].value_counts().sum()
total_missed = out_of_zone[out_of_zone['actual_strike'] != out_of_zone['called_strike']]['PitchCall'].value_counts().sum()
total_accuracy = 1 - (total_missed / total_called)

worksheet['G42'] = total_called
worksheet['H42'] = total_missed
worksheet['I42'] = f'{round(total_accuracy * 100, 1)}%'
worksheet['I42'].alignment = openpyxl.styles.Alignment(horizontal='right')


workbook.save('ump_scorecards/' + title)